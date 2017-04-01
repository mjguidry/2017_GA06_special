# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 21:14:30 2017

@author: mike
"""


from mpl_toolkits.basemap import Basemap
import pylab as plt
import string
from scipy.spatial import ConvexHull
import numpy as np
import pickle
import os.path
from openpyxl import load_workbook, Workbook

def area_for_polygon(polygon):
    result = 0
    imax = len(polygon) - 1
    for i in range(0,imax):
        result += (polygon[i][0] * polygon[i+1][1]) - (polygon[i+1][0] * polygon[i][1])
    result += (polygon[imax][0] * polygon[0][1]) - (polygon[0][0] * polygon[imax][1])
    return result / 2.

def centroid_for_polygon(polygon):
    area = area_for_polygon(polygon)
    imax = len(polygon) - 1

    result_x = 0
    result_y = 0
    for i in range(0,imax):
        result_x += (polygon[i][0] + polygon[i+1][0]) * ((polygon[i][0] * polygon[i+1][1]) - (polygon[i+1][0] * polygon[i][1]))
        result_y += (polygon[i][1] + polygon[i+1][1]) * ((polygon[i][0] * polygon[i+1][1]) - (polygon[i+1][0] * polygon[i][1]))
    result_x += (polygon[imax][0] + polygon[0][0]) * ((polygon[imax][0] * polygon[0][1]) - (polygon[0][0] * polygon[imax][1]))
    result_y += (polygon[imax][1] + polygon[0][1]) * ((polygon[imax][0] * polygon[0][1]) - (polygon[0][0] * polygon[imax][1]))
    result_x /= (area * 6.0)
    result_y /= (area * 6.0)

    return [result_x,result_y]

fig = plt.figure(1,figsize=(20,10))
plt.clf()

m=Basemap(projection='mill',
           urcrnrlat=34.186289,
           urcrnrlon=-84.097692,
           llcrnrlat=33.833294,
           llcrnrlon=-84.571683)
m.readshapefile('./GA06/GA06', 
                 'GA06', 
                 drawbounds = False)                 



maxx=0
maxy=0
minx=1e30
miny=1e30
centroids=dict()
precincts=dict()
for info, shape in zip(m.GA06_info, m.GA06):
    county=info['County_12']
    precinct=info['Precinct_1']
#    if(county=='Fulton'):
#        precinct=string.upper(precinct)
#    else:
#        precinct=string.capwords(precinct)
#    name=county+' - '+precinct
    name=precinct
    precincts[precinct]=county
    x, y = zip(*shape)
    hull=ConvexHull(shape)
    centroid=centroid_for_polygon(np.array(shape)[hull.vertices])
    lon,lat=m(centroid[0],centroid[1],inverse=True)
    centroids[name]=[lon,lat]
    m.plot(x, y, marker=None,color='w')
    m.plot(centroid[0],centroid[1],marker='o',color='y')
    plt.fill(x,y,'b')
    xs=[elem[0] for elem in shape]
    ys=[elem[1] for elem in shape]
    maxx=max(max(xs),maxx)
    maxy=max(max(ys),maxy)
    minx=min(min(xs),minx)
    miny=min(min(ys),miny)

print m(maxx, maxy,inverse=True)
print m(minx, miny,inverse=True) 

f=open('centroids.pkl','wb')
pickle.dump(centroids,f)
f.close()

if(os.path.exists('GA06_blocks.xlsx')):
    wb=load_workbook('GA06_blocks.xlsx')
else:
    wb=Workbook()
sheets=wb.get_sheet_names()
if('Precincts' in sheets):
    ws=wb.get_sheet_by_name('Precincts')
else:
    ws=wb.create_sheet('Precincts')

counties=sorted(set(precincts.values()))
row=2
for county in counties:
    keys=sorted([precinct for precinct in precincts.keys() if precincts[precinct]==county])
    for key in keys:
        ws.cell(row=row,column=6).value=county
        ws.cell(row=row,column=7).value=key
        row=row+1

wb.save('GA06_blocks.xlsx')
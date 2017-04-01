# -*- coding: utf-8 -*-
"""
Created on Sat Apr 01 12:51:35 2017

@author: MGuidry
"""

from random import shuffle, sample
from math import sqrt
import numpy as np
import colorsys
from openpyxl import load_workbook
import pickle
import re
from numpy import array, zeros
from openpyxl import Workbook
from openpyxl.utils import _get_column_letter
from openpyxl.styles import Font, PatternFill,colors
from openpyxl.comments import Comment

execfile('_hungarian.py')

def _get_colors(num_colors):
    colors=[]
    for i in np.arange(0., 360., 360. / num_colors):
        hue = i/360.
        lightness = (50 + np.random.rand() * 10)/100.
        saturation = (90 + np.random.rand() * 10)/100.
        (r,g,b)=colorsys.hls_to_rgb(hue, lightness, saturation)
        colors.append('%02x%02x%02x' % (int(r*256), int(g*256), int(b*256)))
    return colors
    
def dist_err(precinct,cell,gray_dict,centroids):
    xmin=min([centroids[x][0] for x in centroids])
    xmax=max([centroids[x][0] for x in centroids])
    ymin=min([centroids[x][1] for x in centroids])
    ymax=max([centroids[x][1] for x in centroids])
    bbox=[xmin, ymin, xmax, ymax]
    min_x=min([x[1] for x in gray_dict])
    max_x=max([x[1] for x in gray_dict])+1
    min_y=min([x[0] for x in gray_dict])
    max_y=max([x[0] for x in gray_dict])+1
    m_x=(bbox[2]-bbox[0])/(max_x-min_x)
    b_x=bbox[2]-m_x*max_x
    m_y=(bbox[3]-bbox[1])/(min_y-max_y)
    b_y=bbox[3]-m_y*min_y    
    x0=centroids[precinct][0]
    y0=centroids[precinct][1]
    x1=m_x*(cell[1]+0.5)+b_x
    y1=m_y*(cell[0]+0.5)+b_y
    #print x0,y0,x1,y1
    tot_err=sqrt((x0-x1)**2+(y0-y1)**2)
    return tot_err**4

def isAdjacent(state,gray_dict):
    blocks=[x for x in gray_dict if state in gray_dict[x]]
    return_val=True
    if (len(blocks)==1):
        return_val=True
    else:
        for block in blocks:
            if((block[0]-1,block[1]) in blocks or
               (block[0]+1,block[1]) in blocks or
               (block[0],block[1]-1) in blocks or
               (block[0],block[1]+1) in blocks):
                pass
            else:
                return_val=False
    return return_val

# First, get outline
gray_dict=dict()
wb_trace=load_workbook('GA06_trace.xlsx',read_only=True)
ws=wb_trace.active
min_row=7
max_row=49
min_col=32
max_col=80
gray='FF808080'
for row in range(min_row,max_row+1):
    for col in range(min_col,max_col+1):
        cell=ws.cell(row=row,column=col)
        color=cell.style.fill.start_color.index
        if(color==gray):
            gray_dict[(row-min_row+1,col-min_col+1)]=''

#Second, get county geographic data
pkl_file=open('centroids.pkl','rb')
centroids=pickle.load(pkl_file)
pkl_file.close()
precinct_keys=sorted(centroids.keys())
cells=sorted(gray_dict.keys())

precinct_arr=[]
#Third, get block data
wb=load_workbook('GA06_blocks.xlsx', read_only=True, data_only=True)
ws=wb.get_sheet_by_name('Precincts')
for row in range(2,210):
    precinct=ws.cell(row=row,column=7).value
    count=int(ws.cell(row=row,column=10).value)
    precinct_arr.extend([precinct]*count)

precincts=sorted(set(precinct_arr))
county_colors=_get_colors(len(precincts))
color_dict=dict()
for k,precinct in enumerate(precincts):
    color_dict[precinct]=county_colors[k]

precinct_keys=sorted(centroids.keys())

#Optimize!
print "Computing Cost Array"
cost_arr=zeros((len(precinct_arr),len(precinct_arr)))
for k,precinct in enumerate(precinct_arr):
    for l,cell in enumerate(cells):
        cost_arr[k,l]=dist_err(precinct,cell,gray_dict,centroids)

print "Computing Best mapping"
best_arr=linear_sum_assignment(cost_arr)
for k in range(len(best_arr[0])):
    precinct=precinct_arr[best_arr[0][k]]
    gray_dict[cells[best_arr[1][k]]]=precinct

adj=[isAdjacent(precinct,gray_dict) for precinct in precincts]
print 'Adjacent precincts = '+str(len([x for x in adj if x==True]))
print 'Non-adjacent precincts = '+str(len([x for x in adj if x==False]))

ft=Font('Arial',size=6)
redFill = PatternFill(start_color=colors.RED,
                   end_color=colors.RED,
                   fill_type='solid')
wb=Workbook()
ws1=wb.active
for cell in gray_dict:
    ws1.cell(row=cell[0],column=cell[1]).value=gray_dict[cell]
    comment = Comment(gray_dict[cell], 'Mike Guidry')
    ws1.cell(row=cell[0],column=cell[1]).comment=comment
    ws1.cell(row=cell[0],column=cell[1]).font=ft
    precinct=gray_dict[cell]
    fill = PatternFill(start_color=color_dict[precinct],
                       end_color=color_dict[precinct],
                       fill_type='solid')
    ws1.cell(row=cell[0],column=cell[1]).fill=fill
    ws1.column_dimensions[_get_column_letter(cell[1])].width=2.8
    ws1.row_dimensions[cell[0]].height=20
    
wb.save(r'GA06_block_map.xlsx')
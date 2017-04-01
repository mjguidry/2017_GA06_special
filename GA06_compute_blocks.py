# -*- coding: utf-8 -*-
"""
Created on Sat Apr 01 09:13:42 2017

@author: MGuidry
"""
import xml.etree.ElementTree as ET
import os.path
from openpyxl import Workbook,load_workbook

xml_files=['./XML/2014_Cobb_detail.xml',
           './XML/2014_DeKalb_detail.xml',
           './XML/2014_Fulton_detail.xml',
           './XML/2016_Cobb_detail.xml',
           './XML/2016_DeKalb_detail.xml',
           './XML/2016_Fulton_detail.xml',
           ]

county_dict=dict()

for f in xml_files:
    tree=ET.parse(f)
    county=tree.find('Region').text
    if(county not in county_dict):
        county_dict[county]=dict()
    date=tree.find('ElectionDate')
    year=int(date.text[-4:])
    contest=[contest for contest in tree.findall('Contest') if contest.attrib['text']=='U.S. Representative, District 6'][0]
    choices=contest.findall('Choice')
    for choice in choices:
        vts=choice.findall('VoteType')
        for vt in vts:
            precincts=vt.getchildren()
            for precinct in precincts:
                votes=int(precinct.attrib['votes'])
                name=precinct.attrib['name']
                if(name not in county_dict[county]):
                    county_dict[county][name]=dict()
                if(year not in county_dict[county][name]):
                    county_dict[county][name][year]=0
                county_dict[county][name][year]=county_dict[county][name][year]+votes

if(os.path.exists('GA06_blocks.xlsx')):
    wb=load_workbook('GA06_blocks.xlsx')
else:
    wb=Workbook()
sheets=wb.get_sheet_names()
if('Precincts' in sheets):
    ws=wb.get_sheet_by_name('Precincts')
else:
    ws=wb.create_sheet('Precincts')

ws.cell(row=1,column=1).value='County'
ws.cell(row=1,column=2).value='Precinct'
ws.cell(row=1,column=3).value='2014'
ws.cell(row=1,column=4).value='2016'

for column in range(1,5):
    cell=ws.cell(row=1,column=column)
    cell.font=cell.font.copy(bold=True)
    
row=2
for county in sorted(county_dict.keys()):
    for precinct in sorted(county_dict[county].keys()):
        ws.cell(row=row,column=1).value=county
        ws.cell(row=row,column=2).value=precinct
        ws.cell(row=row,column=3).value=county_dict[county][precinct].get(2014,0)
        ws.cell(row=row,column=4).value=county_dict[county][precinct].get(2016,0)
        row=row+1

wb.save('GA06_blocks.xlsx')
